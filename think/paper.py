from openpyxl import load_workbook
def Question(que_type,numbers):
    '''
    que_type:试题类型"单选题","多选题","判断题","填空题"
    numbers:需要抽取的试题编号
    '''
    questions = [] #储存抽取的题目
    wb = load_workbook("dataset/题库.xlsx") #载入题库
    if que_type == "单选题":
        ws = wb[que_type]        
        for i in numbers: #按随机生成的编号抽题
            question = ws["B"+str(i)].value #问题在B列
            answerA = "A:\t" + str(ws["C"+str(i)].value) #选项A在C列，"\t"相当于按一下tab键，在字符间产生间隔
            answerB = "B:\t" + str(ws["D"+str(i)].value) #选项B在D列
            answerC = "C:\t" + str(ws["E"+str(i)].value) #选项C在E列
            answerD = "D:\t" + str(ws["F"+str(i)].value) #选项D在F列
            right_answer  = ws["G"+str(i)].value #正确答案在G列
            single_question = [question, answerA, answerB, answerC, answerD, right_answer] #每行的数据存入列表
            questions.append(single_question) #每个题目的数据存入总列表
    elif que_type == "多选题":
        ws = wb[que_type]        
        for i in numbers:
            question = ws["B"+str(i)].value            
            answerA = "A:\t" + str(ws["C"+str(i)].value)
            answerB = "B:\t" + str(ws["D"+str(i)].value)
            answerC = "C:\t" + str(ws["E"+str(i)].value)
            answerD = "D:\t" + str(ws["F"+str(i)].value)
            right_answer  = ws["H"+str(i)].value
            single_question = [question, answerA, answerB, answerC, answerD, right_answer]
            if ws["G"+str(i)].value: #有些题有E选项，有些没有，因此需要判断一下是否有E选项
                answerE = "E:\t" + str(ws["G"+str(i)].value)
                single_question.insert(-1,answerE) #将E选项插入到答案前面，保持答案是最后一个元素
            questions.append(single_question)
    else: #判断题和填空题，内容只取题干和答案
        ws = wb[que_type]        
        for i in numbers:
            question = ws["B"+str(i)].value
            right_answer  = ws["C"+str(i)].value
            single_question = [question, right_answer]
            questions.append(single_question)
            
    return questions

num_single_choice = range(566+1)[2:]
num_mult_choice = range(196+1)[2:]
num_judgment = range(418+1)[2:]
num_completion = range(190+1)[2:]

question_num = {"单选题号":num_single_choice,
           "多选题号":num_mult_choice,
            "判断题号":num_judgment,
            "填空题号":num_completion
    }

questions_data = {
        "单选题":Question("单选题", question_num["单选题号"]),
        "多选题":Question("多选题", question_num["多选题号"]),
        "判断题":Question("判断题", question_num["判断题号"]),
        "填空题":Question("填空题", question_num["填空题号"])
    }

print(questions_data)